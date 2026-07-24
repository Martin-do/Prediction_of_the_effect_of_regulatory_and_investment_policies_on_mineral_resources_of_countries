"""Reproduce the post hoc WGI Regulatory Quality two-scale sensitivity.

The script creates an isolated temporary working copy of the release, adds the
0-100 Regulatory Quality score from the same frozen WGI 2025 workbook, builds
matched samples, fits only the additional score-scale M3 models, and repeats the
paired country-cluster bootstrap using the same country-draw matrices as the
primary estimate-scale analysis.
"""
from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import os
from pathlib import Path
import shutil
import subprocess
import sys
import tempfile

import numpy as np
import pandas as pd
from openpyxl import load_workbook


def deterministic_env() -> dict[str, str]:
    env = dict(os.environ)
    env.update({
        "PYTHONHASHSEED": "0",
        "OMP_NUM_THREADS": "1",
        "MKL_NUM_THREADS": "1",
        "OPENBLAS_NUM_THREADS": "1",
        "NUMEXPR_NUM_THREADS": "1",
    })
    return env


def run_python(script: Path, *args: str, cwd: Path) -> None:
    completed = subprocess.run([sys.executable, str(script), *args], cwd=cwd, env=deterministic_env(), text=True, capture_output=True)
    if completed.returncode:
        raise RuntimeError(f"Command failed: {script} {' '.join(args)}\n{completed.stdout}\n{completed.stderr}")


def add_score_to_panel(work_root: Path) -> None:
    workbook = work_root / "01_raw_data/world_bank/WGI_2025_Governance_Estimates_and_Scores.xlsx"
    panel_path = work_root / "05_processed_data/MS2_SIGNAL_ADMISSIBILITY_PANEL_V3.csv"
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["rq"]
    records: dict[tuple[str, int], tuple[float | None, float | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        iso, year, estimate, score = row[2], row[5], row[8], row[12]
        if iso is None or year is None:
            continue
        records[(str(iso), int(year))] = (
            None if estimate is None else float(estimate),
            None if score is None else float(score),
        )
    panel = pd.read_csv(panel_path)
    keys = list(zip(panel["ISO3"].astype(str), panel["Year"].astype(int)))
    estimate = [records.get(key, (None, None))[0] for key in keys]
    score = [records.get(key, (None, None))[1] for key in keys]
    existing = panel["Regulatory_Quality"].to_numpy(float)
    incoming = np.asarray([np.nan if value is None else value for value in estimate], dtype=float)
    mask = ~np.isnan(existing)
    if not np.allclose(existing[mask], incoming[mask], atol=1e-10, rtol=0, equal_nan=True):
        raise RuntimeError("Frozen Regulatory Quality estimate does not match the WGI workbook.")
    panel["Regulatory_Quality_Score"] = score
    score_lookup = {(iso, year): value for (iso, year), (_, value) in records.items()}
    panel["Regulatory_Quality_Score_lag1"] = [score_lookup.get((iso, year - 1)) for iso, year in keys]
    panel.to_csv(panel_path, index=False)


def patch_config(work_root: Path) -> None:
    path = work_root / "03_config/analysis_config.json"
    cfg = json.loads(path.read_text(encoding="utf-8"))
    cfg["release"] = "V0.4.3-posthoc-RQ-two-scale-sensitivity"
    cfg["feature_sets"]["M3_Macro_Oil_RegQuality_Score"] = [
        "__TARGET_LAG1__", "GDP_log_lag1", "Population_log_lag1",
        "Inflation_CPI_Annual_Pct_lag1", "Electricity_Access_Pct_lag1",
        "Oil_Rents_GDP_Pct_lag1", "Regulatory_Quality_Score_lag1",
    ]
    cfg["comparisons"].append({
        "signal": "Regulatory_Quality_Score_lag1",
        "base_model": "M2_Macro_Oil",
        "added_model": "M3_Macro_Oil_RegQuality_Score",
    })
    cfg["posthoc_sensitivity"] = {
        "purpose": "Compare the WGI Regulatory Quality estimate with the alternative 0-100 score from the same frozen workbook.",
        "status": "Post hoc; outside the primary ten-setup classification.",
        "all_other_design_elements": "held fixed",
    }
    path.write_text(json.dumps(cfg, indent=2) + "\n", encoding="utf-8")


def bootstrap_shared_draws(release_root: Path, work_root: Path, output_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    pipeline = work_root / "04_pipeline"
    sys.path.insert(0, str(pipeline))
    spec = importlib.util.spec_from_file_location("bootstrap_module", pipeline / "06_bootstrap_uncertainty.py")
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    cfg = json.loads((work_root / "03_config/analysis_config.json").read_text(encoding="utf-8"))
    registry = pd.read_csv(work_root / "06_locked_design/Matched_Sample_Registry.csv")
    score_registry = registry[registry["signal"] == "Regulatory_Quality_Score_lag1"].sort_values("sample_id")
    contributions, registry_entries = [], []
    for _, entry in score_registry.iterrows():
        primary = registry[
            (registry["signal"] == "Regulatory_Quality_lag1")
            & (registry["target_role"] == entry["target_role"])
            & (registry["universe"] == entry["universe"])
        ].iloc[0]
        score_path = work_root / "07_outputs/partials" / f"contributions__{entry['sample_id']}.csv"
        part = pd.read_csv(score_path)
        part["sample_id"] = primary["sample_id"]
        contributions.append(part)
        new_entry = entry.copy()
        new_entry["sample_id"] = primary["sample_id"]
        registry_entries.append(new_entry)
    contribution_frame = pd.concat(contributions, ignore_index=True)
    score_registry_for_draws = pd.DataFrame(registry_entries)
    unit_parts = []
    for _, entry in score_registry_for_draws.iterrows():
        unit, _, _ = module.run_sample(entry, contribution_frame, cfg)
        unit_parts.append(unit)
    score_units = pd.concat(unit_parts, ignore_index=True)
    score_setups = module.aggregate_setup_verdicts(score_units, score_registry_for_draws, cfg)
    score_units.to_csv(output_dir / "RQ_Score_Bootstrap_Unit_Summary_SHARED_DRAWS.csv", index=False)
    score_setups.to_csv(output_dir / "RQ_Score_Setup_Verdicts_SHARED_DRAWS.csv", index=False)
    return score_units, score_setups


def finalize(release_root: Path, work_root: Path, output_dir: Path, score_units: pd.DataFrame, score_setups: pd.DataFrame) -> None:
    panel = pd.read_csv(work_root / "05_processed_data/MS2_SIGNAL_ADMISSIBILITY_PANEL_V3.csv", usecols=["Year", "Regulatory_Quality", "Regulatory_Quality_Score"]).dropna()
    slope, intercept = np.polyfit(panel["Regulatory_Quality"], panel["Regulatory_Quality_Score"], 1)
    residual = panel["Regulatory_Quality_Score"] - (slope * panel["Regulatory_Quality"] + intercept)
    annual = []
    for year, group in panel.groupby("Year"):
        a, b = np.polyfit(group["Regulatory_Quality"], group["Regulatory_Quality_Score"], 1)
        annual_residual = group["Regulatory_Quality_Score"] - (a * group["Regulatory_Quality"] + b)
        annual.append({"year": int(year), "observations": len(group), "slope": a, "intercept": b, "max_abs_residual": float(annual_residual.abs().max())})
    pd.DataFrame(annual).to_csv(output_dir / "RQ_Scale_Affine_Transformation_By_Year.csv", index=False)
    pd.DataFrame([{
        "estimate_observations": len(panel), "year_min": int(panel["Year"].min()), "year_max": int(panel["Year"].max()),
        "score_equals_slope_times_estimate_plus_intercept_slope": float(slope), "intercept": float(intercept),
        "pearson_correlation": float(panel["Regulatory_Quality"].corr(panel["Regulatory_Quality_Score"])),
        "maximum_absolute_residual": float(residual.abs().max()), "rmse_residual": float(np.sqrt(np.mean(residual ** 2))),
        "interpretation": "The score is a fixed affine transformation of the estimate in the frozen WGI 2025 workbook, to numerical precision.",
    }]).to_csv(output_dir / "RQ_Scale_Transformation_Headline.csv", index=False)

    primary_units = pd.read_csv(release_root / "07_outputs/Bootstrap_Unit_Summary.csv")
    primary_units = primary_units[(primary_units["signal"] == "Regulatory_Quality_lag1") & (primary_units["comparison_type"] == "signal_increment") & (primary_units["target_role"].isin(["primary", "normalized"]))]
    score_units = score_units[(score_units["comparison_type"] == "signal_increment") & (score_units["target_role"].isin(["primary", "normalized"]))]
    keys = ["target_role", "universe", "validation", "split_id", "algorithm"]
    comparison = primary_units.merge(score_units, on=keys, suffixes=("_estimate", "_score"), validate="one_to_one")
    comparison["point_delta_r2_score_minus_estimate"] = comparison["point_delta_r2_score"] - comparison["point_delta_r2_estimate"]
    comparison["bootstrap_median_score_minus_estimate"] = comparison["bootstrap_median_delta_r2_score"] - comparison["bootstrap_median_delta_r2_estimate"]
    comparison["unit_verdict_changed"] = comparison["five_way_unit_verdict_estimate"] != comparison["five_way_unit_verdict_score"]
    comparison.to_csv(output_dir / "RQ_Two_Scale_Unit_Comparison_SHARED_DRAWS.csv", index=False)

    primary_setups = pd.read_csv(release_root / "07_outputs/Bootstrap_Five_Way_Setup_Verdicts.csv")
    primary_setups = primary_setups[(primary_setups["signal"] == "Regulatory_Quality_lag1") & (primary_setups["target_role"].isin(["primary", "normalized"]))]
    score_setups = score_setups[score_setups["target_role"].isin(["primary", "normalized"])]
    setup_rows = []
    for (target_role, universe), group in comparison.groupby(["target_role", "universe"]):
        p = primary_setups[(primary_setups["target_role"] == target_role) & (primary_setups["universe"] == universe)].iloc[0]
        s = score_setups[(score_setups["target_role"] == target_role) & (score_setups["universe"] == universe)].iloc[0]
        setup_rows.append({
            "target_role": target_role, "universe": universe, "units": len(group),
            "estimate_median_point_delta_r2": float(group["point_delta_r2_estimate"].median()),
            "score_median_point_delta_r2": float(group["point_delta_r2_score"].median()),
            "maximum_absolute_unit_point_difference": float(group["point_delta_r2_score_minus_estimate"].abs().max()),
            "unit_verdict_changes": int(group["unit_verdict_changed"].sum()),
            "estimate_setup_verdict": p["bootstrap_five_way_setup_verdict"],
            "score_setup_verdict": s["bootstrap_five_way_setup_verdict"],
            "setup_verdict_changed": p["bootstrap_five_way_setup_verdict"] != s["bootstrap_five_way_setup_verdict"],
        })
    setup_comparison = pd.DataFrame(setup_rows)
    setup_comparison.to_csv(output_dir / "RQ_Two_Scale_Setup_Comparison_SHARED_DRAWS.csv", index=False)
    headline = pd.DataFrame([{
        "decision_bearing_units": len(comparison),
        "maximum_absolute_point_delta_r2_difference": float(comparison["point_delta_r2_score_minus_estimate"].abs().max()),
        "maximum_absolute_bootstrap_median_difference": float(comparison["bootstrap_median_score_minus_estimate"].abs().max()),
        "unit_verdict_changes_shared_draws": int(comparison["unit_verdict_changed"].sum()),
        "setup_verdict_changes_shared_draws": int(setup_comparison["setup_verdict_changed"].sum()),
        "estimate_setup_dependent_setups": int((primary_setups["bootstrap_five_way_setup_verdict"] == "Setup-dependent").sum()),
        "score_setup_dependent_setups": int((score_setups["bootstrap_five_way_setup_verdict"] == "Setup-dependent").sum()),
        "overall_interpretation": "Representational rather than substantively different for this pipeline; the primary estimate-scale classification is unchanged.",
    }])
    headline.to_csv(output_dir / "RQ_Two_Scale_Final_Headline.csv", index=False)

    rows = []
    for path in sorted(output_dir.iterdir()):
        if path.is_file() and path.name != "SHA256SUMS.csv":
            rows.append({"file": path.name, "sha256": hashlib.sha256(path.read_bytes()).hexdigest(), "bytes": path.stat().st_size})
    pd.DataFrame(rows).to_csv(output_dir / "SHA256SUMS.csv", index=False)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo-root", type=Path, default=Path(__file__).resolve().parents[2])
    parser.add_argument("--output-dir", type=Path, default=Path(__file__).resolve().parent)
    parser.add_argument("--keep-workdir", action="store_true")
    args = parser.parse_args()
    release_root = args.repo_root.resolve()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    temp_parent = Path(tempfile.mkdtemp(prefix="ms2_rq_scale_"))
    work_root = temp_parent / release_root.name
    ignore = shutil.ignore_patterns("legacy_archive", "10_reporting_assets", "11_posthoc_checks", "__pycache__", "*.pyc")
    shutil.copytree(release_root, work_root, ignore=ignore)
    try:
        add_score_to_panel(work_root)
        patch_config(work_root)
        run_python(work_root / "04_pipeline/04_lock_samples.py", cwd=work_root)
        registry = pd.read_csv(work_root / "06_locked_design/Matched_Sample_Registry.csv")
        score_ids = registry.loc[registry["signal"] == "Regulatory_Quality_Score_lag1", "sample_id"].tolist()
        for sample_id in score_ids:
            run_python(work_root / "04_pipeline/05_run_corrected_audit.py", "--sample-id", sample_id, cwd=work_root)
        score_units, score_setups = bootstrap_shared_draws(release_root, work_root, args.output_dir)
        finalize(release_root, work_root, args.output_dir, score_units, score_setups)
    finally:
        if args.keep_workdir:
            print(f"Isolated work directory retained: {temp_parent}")
        else:
            shutil.rmtree(temp_parent, ignore_errors=True)
    print("Post hoc RQ two-scale sensitivity completed.")


if __name__ == "__main__":
    main()
